import openpyxl
from openpyxl.drawing.image import Image as XLImage
import pandas as pd
from PIL import Image
import io
import base64
import logging
from typing import List, Dict, Tuple, Optional
import os
from dataclasses import dataclass
import json
from concurrent.futures import ThreadPoolExecutor, as_completed
import hashlib
import time
from enum import Enum
from transformers import pipeline
import torch

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('content_moderation.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class ContentSeverity(Enum):
    SAFE = "safe"
    QUESTIONABLE = "questionable"
    EXPLICIT = "explicit"
    VIOLENT = "violent"
    DISTURBING = "disturbing"

@dataclass
class ModerationResult:
    is_explicit: bool
    confidence: float
    categories: List[str]
    severity: ContentSeverity
    details: Dict

class ImageExtractor:
    """Extract images from Excel cells"""
    
    @staticmethod
    def extract_images_from_xlsx(file_path: str) -> Dict[str, List[Tuple[int, int, bytes]]]:
        """
        Extract all images from Excel file
        Returns: Dict mapping sheet names to list of (row, col, image_bytes)
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True, keep_vba=True)
            images_data = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_images = []
                
                # Handle embedded images
                if hasattr(sheet, '_images'):
                    for image in sheet._images:
                        # Get image position
                        row = image.anchor._from.row
                        col = image.anchor._from.col
                        
                        # Extract image data
                        img_bytes = image.ref
                        sheet_images.append((row, col, img_bytes))
                
                # Handle images in cells (base64 encoded or binary)
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, bytes):
                            sheet_images.append((cell.row, cell.column, cell.value))
                        elif cell.value and isinstance(cell.value, str):
                            # Check if it's base64 encoded image
                            if cell.value.startswith('data:image'):
                                try:
                                    # Extract base64 data
                                    base64_data = cell.value.split(',')[1]
                                    img_bytes = base64.b64decode(base64_data)
                                    sheet_images.append((cell.row, cell.column, img_bytes))
                                except Exception as e:
                                    logger.warning(f"Failed to decode base64 image at {cell.coordinate}: {e}")
                
                if sheet_images:
                    images_data[sheet_name] = sheet_images
            
            return images_data
            
        except Exception as e:
            logger.error(f"Error extracting images from {file_path}: {e}")
            raise

class ContentModerationAPI:
    """Local content moderation using pre-trained ML model (no API keys required)"""
    
    def __init__(self):
        self.rate_limit_delay = 0.1  # Minimal delay for local processing
        self.device = "cuda" if torch.cuda.is_available() else "cpu"
        logger.info(f"Using device: {self.device}")
        
        # Load the NSFW detection pipeline (downloads model on first run)
        self.detector = pipeline(
            "image-classification", 
            model="Falconsai/nsfw_image_detection",
            device=self.device if self.device == "cuda" else -1  # -1 for CPU
        )
        logger.info("NSFW detection model loaded successfully")
    
    def analyze_image(self, image_bytes: bytes) -> ModerationResult:
        """
        Analyze image for explicit content using local model
        """
        try:
            return self._local_nsfw_detection(image_bytes)
                
        except Exception as e:
            logger.error(f"Local model analysis failed: {e}")
            # Return safe by default on failure
            return ModerationResult(
                is_explicit=False,
                confidence=0.0,
                categories=[],
                severity=ContentSeverity.SAFE,
                details={"error": str(e)}
            )
    
    def _local_nsfw_detection(self, image_bytes: bytes) -> ModerationResult:
        """Local NSFW detection using pre-trained model"""
        # Convert bytes to PIL Image
        img = Image.open(io.BytesIO(image_bytes))
        if img.mode != 'RGB':
            img = img.convert('RGB')  # Ensure RGB for model compatibility
        
        time.sleep(self.rate_limit_delay)  # Minimal delay
        
        # Run detection
        prediction = self.detector(img)
        
        # Parse prediction (model outputs 'nsfw' or 'normal')
        top_prediction = prediction[0]
        top_label = top_prediction['label'].lower()
        confidence = top_prediction['score']
        
        is_explicit = top_label == 'nsfw' and confidence > 0.5  # Adjustable threshold
        categories = ['adult'] if top_label == 'nsfw' else []
        
        severity = ContentSeverity.SAFE
        if is_explicit:
            severity = ContentSeverity.EXPLICIT
        elif top_label == 'nsfw' and confidence > 0.3:  # Lower threshold for questionable
            severity = ContentSeverity.QUESTIONABLE
        
        return ModerationResult(
            is_explicit=is_explicit,
            confidence=confidence,
            categories=categories,
            severity=severity,
            details={
                "top_label": top_label,
                "full_prediction": prediction
            }
        )

class ExcelContentModerator:
    """Main class for moderating Excel content"""
    
    def __init__(self, 
                 sensitivity_threshold: float = 0.5,
                 backup_original: bool = True):
        self.extractor = ImageExtractor()
        self.moderator = ContentModerationAPI()  # No API key needed
        self.sensitivity_threshold = sensitivity_threshold
        self.backup_original = backup_original
        self.processed_images_cache = {}  # Cache to avoid reprocessing same images
        
    def process_excel_file(self, input_path: str, output_path: str = None,
                          target_columns: List[str] = None) -> Dict:
        """
        Process Excel file and remove rows with explicit content
        """
        if not output_path:
            base, ext = os.path.splitext(input_path)
            output_path = f"{base}_cleaned{ext}"
            
        # Create backup if requested
        if self.backup_original:
            backup_path = f"{input_path}.backup"
            import shutil
            shutil.copy2(input_path, backup_path)
            logger.info(f"Created backup at {backup_path}")
            
        results = {
            'total_rows': 0,
            'removed_rows': 0,
            'processed_images': 0,
            'explicit_images': 0,
            'errors': [],
            'removed_row_details': []
        }
        
        try:
            # Extract images
            logger.info(f"Extracting images from {input_path}")
            images_data = self.extractor.extract_images_from_xlsx(input_path)
            
            # Load Excel with pandas for easier row manipulation
            excel_file = pd.ExcelFile(input_path)
            output_sheets = {}
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                results['total_rows'] += len(df)
                
                if sheet_name in images_data:
                    rows_to_remove = set()
                    
                    # Process images with threading for performance
                    with ThreadPoolExecutor(max_workers=5) as executor:
                        future_to_image = {}
                        
                        for row_idx, col_idx, img_bytes in images_data[sheet_name]:
                            # Create hash for caching
                            img_hash = hashlib.md5(img_bytes).hexdigest()
                            
                            if img_hash in self.processed_images_cache:
                                # Use cached result
                                moderation_result = self.processed_images_cache[img_hash]
                            else:
                                # Submit for processing
                                future = executor.submit(self.moderator.analyze_image, img_bytes)
                                future_to_image[future] = (row_idx, col_idx, img_hash)
                        
                        # Collect results
                        for future in as_completed(future_to_image):
                            row_idx, col_idx, img_hash = future_to_image[future]
                            
                            try:
                                moderation_result = future.result(timeout=30)
                                self.processed_images_cache[img_hash] = moderation_result
                                
                                results['processed_images'] += 1
                                
                                if moderation_result.is_explicit or \
                                   (moderation_result.confidence > self.sensitivity_threshold and 
                                    moderation_result.severity != ContentSeverity.SAFE):
                                    
                                    rows_to_remove.add(row_idx - 1)  # Adjust for 0-based pandas index
                                    results['explicit_images'] += 1
                                    results['removed_row_details'].append({
                                        'sheet': sheet_name,
                                        'row': row_idx,
                                        'column': col_idx,
                                        'severity': moderation_result.severity.value,
                                        'confidence': moderation_result.confidence,
                                        'categories': moderation_result.categories
                                    })
                                    
                                    logger.warning(f"Explicit content found in {sheet_name} at row {row_idx}, col {col_idx}")
                                    
                            except Exception as e:
                                logger.error(f"Error processing image at {sheet_name}[{row_idx},{col_idx}]: {e}")
                                results['errors'].append(str(e))
                    
                    # Remove flagged rows
                    if rows_to_remove:
                        df_cleaned = df.drop(index=list(rows_to_remove))
                        results['removed_rows'] += len(rows_to_remove)
                        logger.info(f"Removed {len(rows_to_remove)} rows from {sheet_name}")
                    else:
                        df_cleaned = df
                else:
                    df_cleaned = df
                    
                output_sheets[sheet_name] = df_cleaned
            
            # Save cleaned Excel file
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, df in output_sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
            logger.info(f"Cleaned file saved to {output_path}")
            
            # Save detailed report
            report_path = f"{output_path}_report.json"
            with open(report_path, 'w') as f:
                json.dump(results, f, indent=2)
            logger.info(f"Report saved to {report_path}")
            
            return results
            
        except Exception as e:
            logger.error(f"Critical error processing file: {e}")
            results['errors'].append(f"Critical error: {e}")
            raise

# Main execution
if __name__ == "__main__":
    # Configuration (no API keys needed)
    SENSITIVITY = 0.5  # 0.0 to 1.0 - threshold for flagging NSFW
    
    # Initialize moderator
    moderator = ExcelContentModerator(
        sensitivity_threshold=SENSITIVITY,
        backup_original=True
    )
    
    # Process file
    input_file = "data_with_images.xlsx"
    output_file = "data_cleaned.xlsx"
    
    try:
        results = moderator.process_excel_file(
            input_path=input_file,
            output_path=output_file,
            target_columns=None  # Process all columns, or specify like ['ImageColumn1', 'ImageColumn2']
        )
        
        print("\n=== Processing Complete ===")
        print(f"Total rows processed: {results['total_rows']}")
        print(f"Rows removed: {results['removed_rows']}")
        print(f"Images analyzed: {results['processed_images']}")
        print(f"Explicit images found: {results['explicit_images']}")
        
        if results['errors']:
            print(f"\nErrors encountered: {len(results['errors'])}")
            for error in results['errors'][:5]:  # Show first 5 errors
                print(f"  - {error}")
                
    except Exception as e:
        logger.error(f"Failed to process file: {e}")
        raise