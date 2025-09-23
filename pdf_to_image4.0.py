import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import logging
from pathlib import Path

# Set up logging
logging.basicConfig(level=logging.INFO,
                   format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def add_images_to_excel(excel_file, images_folder, output_excel=None):
    """
    Add images to Excel file based on matching filenames in the "File Name" column.
    
    Args:
        excel_file (str): Path to the Excel file
        images_folder (str): Path to the folder containing images
        output_excel (str): Path for the output Excel file. If None, will create 'updated_' prefix
    """
    try:
        # Load the workbook
        wb = load_workbook(excel_file)
        ws = wb.active
        logger.info(f"Loaded Excel file: {excel_file}")

        # Find the "File Name" column
        filename_col = None
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header and "file name" in str(header).lower():
                filename_col = col
                break

        if not filename_col:
            logger.error('Column "File Name" not found in Excel file')
            return

        # Get the next empty column for images
        image_col = ws.max_column + 1
        
        # Add header for image column
        ws.cell(row=1, column=image_col, value="Screenshot")
        
        # Set column width for images
        ws.column_dimensions[chr(64 + image_col)].width = 40  # Excel columns start at A(65)

        # Process each row starting from row 2 (after header)
        images_added = 0
        for row in range(2, ws.max_row + 1):
            filename = ws.cell(row=row, column=filename_col).value
            if not filename:
                continue

            # Clean filename and try different variations
            filename = str(filename).strip()
            filename_no_ext = os.path.splitext(filename)[0]
            
            # Look for matching image
            image_path = None
            for img_file in os.listdir(images_folder):
                img_name_no_ext = os.path.splitext(img_file)[0]
                # Check if filename matches either exactly or without extension
                # Also handle single-page PDFs (no _page_1 suffix)
                if (img_name_no_ext == filename_no_ext or
                    img_name_no_ext == f"{filename_no_ext}_page_1" or
                    img_file == filename):
                    image_path = os.path.join(images_folder, img_file)
                    break

            if image_path:
                try:
                    # Add image to worksheet
                    img = Image(image_path)
                    
                    # Resize image
                    img.width = 200
                    img.height = 200
                    
                    # Calculate cell position
                    cell = f"{chr(64 + image_col)}{row}"
                    
                    # Add image
                    ws.add_image(img, cell)
                    
                    # Adjust row height
                    ws.row_dimensions[row].height = 150
                    
                    images_added += 1
                    logger.info(f"Added image for {filename} in row {row}")
                except Exception as e:
                    logger.error(f"Error adding image for {filename}: {str(e)}")
            else:
                logger.warning(f"No matching image found for {filename}")

        # Save the workbook
        if not output_excel:
            output_excel = f"updated_{os.path.basename(excel_file)}"
        wb.save(output_excel)
        logger.info(f"Saved updated Excel file as {output_excel}")
        logger.info(f"Total images added: {images_added}")

    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")

if __name__ == "__main__":
    # Get the directory where this script is located
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Define paths
    excel_file = r"C:\Users\sk23dg\Desktop\Labins_Workflow\PDF_Extractor\pdf_extractor_Draft\output3.0.xlsx"  # Update this path
    images_folder = os.path.join(script_dir, "extracted_screenshots")  # Folder with extracted images
    output_excel = "updated_excel_with_screenshots.xlsx"
    
    # Add images to Excel
    add_images_to_excel(excel_file, images_folder, output_excel)