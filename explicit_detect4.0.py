import openpyxl
from openpyxl.drawing.image import Image as XLImage
import pandas as pd
from PIL import Image
import io
import base64
import logging
from typing import List, Dict, Tuple, Optional
import os
import glob
from dataclasses import dataclass
import json
from concurrent.futures import ThreadPoolExecutor, as_completed
import hashlib
import time
from enum import Enum
from transformers import pipeline
import torch
import argparse
import shutil

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('content_moderation.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class ContentSeverity(Enum):
    SAFE = "safe"
    QUESTIONABLE = "questionable"
    EXPLICIT = "explicit"
    VIOLENT = "violent"
    DISTURBING = "disturbing"

@dataclass
class ModerationResult:
    is_explicit: bool
    confidence: float
    categories: List[str]
    severity: ContentSeverity
    details: Dict


class ImageExtractor:
    """Extract images from Excel cells"""
    
    @staticmethod
    def extract_images_from_xlsx(file_path: str) -> Dict[str, List[Tuple[int, int, bytes]]]:
        """
        Extract all images from Excel file
        Returns: Dict mapping sheet names to list of (row, col, image_bytes)
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True, keep_vba=True)
            images_data = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_images = []
                
                # Handle embedded images
                if hasattr(sheet, '_images'):
                    for image in sheet._images:
                        # Get image position
                        row = image.anchor._from.row
                        col = image.anchor._from.col
                        
                        # Extract image data
                        img_bytes = image.ref
                        sheet_images.append((row, col, img_bytes))
                
                # Handle images in cells (base64 encoded or binary)
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, bytes):
                            sheet_images.append((cell.row, cell.column, cell.value))
                        elif cell.value and isinstance(cell.value, str):
                            # Check if it's base64 encoded image
                            if cell.value.startswith('data:image'):
                                try:
                                    # Extract base64 data
                                    base64_data = cell.value.split(',')[1]
                                    img_bytes = base64.b64decode(base64_data)
                                    sheet_images.append((cell.row, cell.column, img_bytes))
                                except Exception as e:
                                    logger.warning(f"Failed to decode base64 image at {cell.coordinate}: {e}")
                
                if sheet_images:
                    images_data[sheet_name] = sheet_images
            
            return images_data
        
        except Exception as e:
            logger.error(f"Error extracting images from {file_path}: {e}")
            raise


class ContentModerationAPI:
    """Local content moderation using pre-trained ML model (no API keys required)"""
    
    def __init__(self):
        self.rate_limit_delay = 0.1  # Minimal delay for local processing
        self.device = "cuda" if torch.cuda.is_available() else "cpu"
        logger.info(f"Using device: {self.device}")
        
        # Load the NSFW detection pipeline (downloads model on first run)
        self.detector = pipeline(
            "image-classification",
            model="Falconsai/nsfw_image_detection",
            device=self.device if self.device == "cuda" else -1  # -1 for CPU
        )
        logger.info("NSFW detection model loaded successfully")
    
    def analyze_image(self, image_bytes: bytes) -> ModerationResult:
        """
        Analyze image for explicit content using local model
        """
        try:
            return self._local_nsfw_detection(image_bytes)
                
        except Exception as e:
            logger.error(f"Local model analysis failed: {e}")
            # Return safe by default on failure
            return ModerationResult(
                is_explicit=False,
                confidence=0.0,
                categories=[],
                severity=ContentSeverity.SAFE,
                details={"error": str(e)}
            )
    
    def _local_nsfw_detection(self, image_bytes: bytes) -> ModerationResult:
        """Local NSFW detection using pre-trained model"""
        # Convert bytes to PIL Image
        img = Image.open(io.BytesIO(image_bytes))
        if img.mode != 'RGB':
            img = img.convert('RGB')  # Ensure RGB for model compatibility
        
        time.sleep(self.rate_limit_delay)  # Minimal delay
        
        # Run detection
        prediction = self.detector(img)
        
        # Parse prediction (model outputs 'nsfw' or 'normal')
        top_prediction = prediction[0]
        top_label = top_prediction['label'].lower()
        confidence = top_prediction['score']
        
        is_explicit = top_label == 'nsfw' and confidence > 0.5  # Adjustable threshold
        categories = ['adult'] if top_label == 'nsfw' else []
        
        severity = ContentSeverity.SAFE
        if is_explicit:
            severity = ContentSeverity.EXPLICIT
        elif top_label == 'nsfw' and confidence > 0.3:  # Lower threshold for questionable
            severity = ContentSeverity.QUESTIONABLE
        
        return ModerationResult(
            is_explicit=is_explicit,
            confidence=confidence,
            categories=categories,
            severity=severity,
            details={
                "top_label": top_label,
                "full_prediction": prediction
            }
        )


class ExcelContentModerator:
    """Main class for moderating Excel content"""
    
    def __init__(self, 
                 sensitivity_threshold: float = 0.5,
                 backup_original: bool = True):
        self.extractor = ImageExtractor()
        self.moderator = ContentModerationAPI()  # No API key needed
        self.sensitivity_threshold = sensitivity_threshold
        self.backup_original = backup_original
        self.processed_images_cache = {}  # Cache to avoid reprocessing same images
        
    def process_excel_file(self, input_path: str, output_path: str = None,
                      target_columns: List[str] = None) -> Dict:
        """
        Process a single Excel file and remove rows with explicit content
        """
        if not output_path:
            base, ext = os.path.splitext(input_path)
            output_path = f"{base}_cleaned{ext}"
        
        # Create backup if requested
        if self.backup_original:
            backup_path = f"{input_path}.backup"
            shutil.copy2(input_path, backup_path)
            logger.info(f"Created backup at {backup_path}")
        
        results = {
            'file': input_path,
            'total_rows': 0,
            'removed_rows': 0,
            'processed_images': 0,
            'explicit_images': 0,
            'errors': [],
            'removed_row_details': []
        }
        
        try:
            # Load Excel with pandas first to get total row count
            excel_file = pd.ExcelFile(input_path)
            output_sheets = {}
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                results['total_rows'] += len(df)
                output_sheets[sheet_name] = df
            
            # Extract images if present
            logger.info(f"Extracting images from {input_path}")
            try:
                images_data = self.extractor.extract_images_from_xlsx(input_path)
                logger.info(f"Found images in sheets: {list(images_data.keys())}")
            except Exception as e:
                logger.warning(f"Error extracting images from {input_path}: {e}")
                images_data = {}
            
            # Process each sheet
            for sheet_name in excel_file.sheet_names:
                df = output_sheets[sheet_name]
                
                if sheet_name in images_data and images_data[sheet_name]:
                    rows_to_remove = set()
                    
                    # Process images with threading for performance
                    with ThreadPoolExecutor(max_workers=5) as executor:
                        future_to_image = {}
                        
                        for row_idx, col_idx, img_bytes in images_data[sheet_name]:
                            # Create hash for caching
                            img_hash = hashlib.md5(img_bytes).hexdigest()
                            
                            if img_hash in self.processed_images_cache:
                                moderation_result = self.processed_images_cache[img_hash]
                                self._handle_moderation_result(
                                    moderation_result, row_idx, col_idx,
                                    sheet_name, rows_to_remove, results
                                )
                            else:
                                future = executor.submit(self.moderator.analyze_image, img_bytes)
                                future_to_image[future] = (row_idx, col_idx, img_hash)
                    
                        # Collect results
                        for future in as_completed(future_to_image):
                            row_idx, col_idx, img_hash = future_to_image[future]
                            try:
                                moderation_result = future.result(timeout=30)
                                self.processed_images_cache[img_hash] = moderation_result
                                results['processed_images'] += 1
                                self._handle_moderation_result(
                                    moderation_result, row_idx, col_idx,
                                    sheet_name, rows_to_remove, results
                                )
                            except Exception as e:
                                logger.error(f"Error processing image at {sheet_name}[{row_idx},{col_idx}]: {e}")
                                results['errors'].append(str(e))
                # End of processing images for sheet

        # Add outer exception handling to close the main try block
        except Exception as e:
            logger.error(f"Critical error processing file: {e}")
            results['errors'].append(f"Critical error: {e}")
            raise

        # Remove rows marked for deletion
        for sheet_name, df in output_sheets.items():
            if sheet_name in images_data:
                # Remove rows with explicit content
                rows_to_keep = df.index.difference(rows_to_remove)
                output_sheets[sheet_name] = df.loc[rows_to_keep]
                results['removed_rows'] += len(df) - len(rows_to_keep)
                logger.info(f"Removed {len(df) - len(rows_to_keep)} explicit rows from {sheet_name}")
        
        # Save cleaned Excel file
        logger.info(f"Saving cleaned Excel file to {output_path}")
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, df in output_sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        logger.info(f"Processing complete: {results}")
        return results
    
    def _handle_moderation_result(self, result: ModerationResult, row_idx: int, col_idx: int,
                                  sheet_name: str, rows_to_remove: set, results: Dict):
        """
        Handle the moderation result for an image
        """
        try:
            if result.is_explicit:
                rows_to_remove.add(row_idx)
                results['explicit_images'] += 1
                results['removed_row_details'].append({
                    'sheet': sheet_name,
                    'row': row_idx,
                    'confidence': result.confidence,
                    'categories': result.categories,
                    'severity': result.severity.value
                })
                logger.info(f"Marked row {row_idx} in {sheet_name} for removal (explicit content)")
            else:
                logger.info(f"Image at {sheet_name}[{row_idx},{col_idx}] is safe")
        
        except Exception as e:
            logger.error(f"Error handling moderation result for {sheet_name}[{row_idx},{col_idx}]: {e}")
            results['errors'].append(str(e))


def main():
    parser = argparse.ArgumentParser(description="Excel Content Moderator")
    parser.add_argument("input", nargs="*", help="Input Excel file(s)")
    parser.add_argument("--input_dir", help="Directory containing input Excel files")
    parser.add_argument("-o", "--output_dir", default=".", help="Output directory for cleaned files and reports")
    parser.add_argument("-t", "--threshold", type=float, default=0.5, help="Detection threshold for explicit content")
    parser.add_argument("--no-backup", action="store_true", help="Do not create backup of original files")
    args = parser.parse_args()

    if args.input_dir:
        input_path = args.input_dir
    elif args.input and len(args.input) > 0:
        if len(args.input) == 1:
            input_path = args.input[0]
        else:
            input_path = args.input  # list of files
    else:
        parser.error("No input provided. Please provide an input file or --input_dir.")

    # Initialize moderator
    moderator = ExcelContentModerator(
        sensitivity_threshold=args.threshold,
        backup_original=not args.no_backup
    )

    # Process each input file or directory
    if os.path.isdir(input_path):
        # Input is a directory, process all Excel files in the directory
        results_summary = {
            'total_files': 0,
            'processed_files': 0,
            'errors': []
        }
        
        for file_name in os.listdir(input_path):
            if file_name.endswith(('.xlsx', '.xls')):
                results_summary['total_files'] += 1
                file_path = os.path.join(input_path, file_name)
                
                try:
                    logger.info(f"Processing file: {file_path}")
                    results = moderator.process_excel_file(file_path)
                    logger.info(f"File processed successfully: {results}")
                    results_summary['processed_files'] += 1
                except Exception as e:
                    logger.error(f"Error processing file {file_path}: {e}")
                    results_summary['errors'].append(str(e))
        
        # Print summary report
        logger.info(f"Batch processing complete: {results_summary}")
        print(f"Processed {results_summary['processed_files']} out of {results_summary['total_files']} files successfully.")
        if results_summary['errors']:
            print("Errors encountered:")
            for error in results_summary['errors']:
                print(f"- {error}")
    else:
        # Single file processing
        try:
            logger.info(f"Processing file: {input_path}")
            results = moderator.process_excel_file(input_path)
            logger.info(f"File processed successfully: {results}")
        except Exception as e:
            logger.error(f"Error processing file {input_path}: {e}")
            print(f"Error: {e}")


if __name__ == "__main__":
    main()
