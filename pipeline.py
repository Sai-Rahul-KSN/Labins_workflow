"""
Complete Pipeline for PDF Data Extraction to Excel
Extracts form data, embedded images, and PDF screenshots into a single Excel file
"""

import sys
import os
import time
import pandas as pd
import re
from datetime import datetime
from pathlib import Path
import subprocess
import argparse
import logging

# Third-party imports
try:
    from PyPDF2 import PdfReader
    from PyPDF2.generic import IndirectObject, DecodedStreamObject
except ImportError:
    print("ERROR: PyPDF2 is required. Install with: pip install PyPDF2")
    sys.exit(1)

try:
    import fitz  # PyMuPDF
except ImportError:
    print("ERROR: PyMuPDF is required. Install with: pip install pymupdf")
    sys.exit(1)

try:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ==================== CONFIGURATION ====================
class Config:
    # Directories
    PDF_DIR = r"C:\Users\sk23dg\Desktop\Content_extraction\test_pdfs"
    OUTPUT_DIR = r"C:\Users\sk23dg\Desktop\Content_extraction\output"
    SCREENSHOT_DIR = None  # Will be set in main
    EMBEDDED_IMG_DIR = None  # Will be set in main
    
    # Output files
    EXCEL_OUTPUT = None  # Will be set in main
    
    # PDF Form Field Names
    DOC_NUM_FIELD = "Doc Num"
    CORNER_OF_SECTION_FIELD = "Corner of Section"
    TOWNSHIP_FIELD = "Township"
    RANGE_FIELD = "Range"
    COUNTY_FIELD = "County"
    SURVEY_IMAGE_FIELD = "Survey Image"
    
    # Image extraction settings
    DPI = 150  # For screenshot quality
    ZOOM = 2  # For PDF to image conversion

# ==================== PDF FORM EXTRACTION ====================
class PDFFormExtractor:
    def __init__(self):
        self.config = Config()
    
    def load_reader(self, pdf_path):
        if not os.path.exists(pdf_path):
            raise FileNotFoundError(f"PDF not found: {pdf_path}")
        return PdfReader(pdf_path)
    
    def get_fields_dict(self, reader):
        try:
            return reader.get_fields()
        except Exception:
            root = reader.trailer.get("/Root")
            acro = root.get("/AcroForm") if root else None
            fields = {}
            if acro and "/Fields" in acro.get_object():
                for f in acro.get_object()["/Fields"]:
                    obj = f.get_object()
                    name = obj.get("/T")
                    if name:
                        fields[name] = obj
            return fields
    
    def find_field_obj(self, reader, field_name):
        root = reader.trailer.get("/Root")
        acro = root.get("/AcroForm") if root else None
        if not acro:
            return None
        for f in acro.get_object().get("/Fields", []):
            obj = f.get_object()
            if obj.get("/T") == field_name:
                return obj
        return None
    
    def get_value(self, fields_dict, field_name):
        try:
            entry = fields_dict.get(field_name)
            if entry is None:
                return ""
            val = entry.get("/V")
            return "" if val is None else str(val)
        except Exception as e:
            logger.warning(f"Could not read value for '{field_name}': {e}")
            return ""
    
    def get_choice_options(self, reader, field_name):
        options = []
        try:
            obj = self.find_field_obj(reader, field_name)
            if not obj:
                return options
            opt = obj.get("/Opt")
            if not opt:
                return options
            for item in opt:
                if isinstance(item, list) and len(item) >= 1:
                    options.append(str(item[0]))
                else:
                    options.append(str(item))
        except Exception as e:
            logger.warning(f"Could not read options for '{field_name}': {e}")
        return options
    
    def detect_image_in_button(self, reader, field_name):
        try:
            obj = self.find_field_obj(reader, field_name)
            if not obj:
                return False
            if obj.get("/FT") != "/Btn":
                return False
            mk = obj.get("/MK")
            if mk and mk.get("/I") is not None:
                return True
            ap = obj.get("/AP")
            if ap and "/N" in ap:
                apn = ap["/N"]
                try:
                    apn_obj = apn.get_object()
                except Exception:
                    apn_obj = apn
                if isinstance(apn_obj, DecodedStreamObject):
                    resources = apn_obj.get("/Resources")
                    if resources:
                        xo = resources.get("/XObject")
                        if isinstance(xo, IndirectObject):
                            xo = xo.get_object()
                        if xo and hasattr(xo, "items"):
                            for _name, xobj in xo.items():
                                try:
                                    xobj_obj = xobj.get_object()
                                    if xobj_obj.get("/Subtype") == "/Image":
                                        return True
                                except Exception:
                                    continue
            return False
        except Exception as e:
            logger.warning(f"Image detection error for '{field_name}': {e}")
            return False
    
    def process_pdf(self, pdf_path):
        logger.info(f"Extracting form data from {pdf_path}")
        try:
            reader = self.load_reader(pdf_path)
            fields = self.get_fields_dict(reader)
            
            # Read field values
            doc_num = self.get_value(fields, self.config.DOC_NUM_FIELD)
            corner_of_section = self.get_value(fields, self.config.CORNER_OF_SECTION_FIELD)
            township = self.get_value(fields, self.config.TOWNSHIP_FIELD)
            range_value = self.get_value(fields, self.config.RANGE_FIELD)
            county = self.get_value(fields, self.config.COUNTY_FIELD)
            
            # Read dropdown options
            township_options = self.get_choice_options(reader, self.config.TOWNSHIP_FIELD)
            range_options = self.get_choice_options(reader, self.config.RANGE_FIELD)
            county_options = self.get_choice_options(reader, self.config.COUNTY_FIELD)
            
            # Detect image presence
            image_present_bool = self.detect_image_in_button(reader, self.config.SURVEY_IMAGE_FIELD)
            image_present_flag = "Y" if image_present_bool else "N"
            
            return {
                "filename": os.path.basename(pdf_path),
                "doc_num": doc_num,
                "corner_of_section": corner_of_section,
                "township": township,
                "range": range_value,
                "county": county,
                "image_present_bool": image_present_bool,
                "image_present_flag": image_present_flag,
                "township_options": ', '.join(township_options) if township_options else '',
                "range_options": ', '.join(range_options) if range_options else '',
                "county_options": ', '.join(county_options) if county_options else ''
            }
        except Exception as e:
            logger.error(f"Error extracting form data from {pdf_path}: {e}")
            return None

# ==================== PDF TO IMAGE CONVERTER ====================
class PDFToImageConverter:
    def __init__(self, output_dir, zoom=2):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.zoom = zoom
    
    def convert_pdf_to_images(self, pdf_path):
        try:
            pdf_name = Path(pdf_path).stem
            logger.info(f"Converting {pdf_path} to screenshots...")
            
            pdf_document = fitz.open(pdf_path)
            image_paths = []
            
            for page_number in range(pdf_document.page_count):
                page = pdf_document[page_number]
                mat = fitz.Matrix(self.zoom, self.zoom)
                pix = page.get_pixmap(matrix=mat)
                
                if pdf_document.page_count == 1:
                    image_path = self.output_dir / f"{pdf_name}.png"
                else:
                    image_path = self.output_dir / f"{pdf_name}_page_{page_number + 1}.png"
                
                pix.save(str(image_path))
                image_paths.append(str(image_path))
                logger.debug(f"Saved screenshot: {image_path}")
            
            pdf_document.close()
            return image_paths
        except Exception as e:
            logger.error(f"Error converting {pdf_path} to images: {e}")
            return []

# ==================== EMBEDDED IMAGE EXTRACTOR ====================
class EmbeddedImageExtractor:
    def __init__(self, output_dir):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def extract_embedded_images(self, pdf_path):
        try:
            pdf_name = Path(pdf_path).stem
            logger.info(f"Extracting embedded images from {pdf_path}")
            
            doc = fitz.open(pdf_path)
            extracted_images = []
            
            for page_index in range(len(doc)):
                page = doc[page_index]
                images = page.get_images(full=True)
                
                for img_index, img in enumerate(images, start=1):
                    try:
                        xref = img[0]
                        base_image = doc.extract_image(xref)
                        image_bytes = base_image.get("image")
                        ext = base_image.get("ext", "png")
                        
                        out_name = f"{pdf_name}_embedded_p{page_index + 1}_img{img_index}.{ext}"
                        out_path = self.output_dir / out_name
                        
                        with open(out_path, "wb") as f:
                            f.write(image_bytes)
                        
                        extracted_images.append(str(out_path))
                        logger.debug(f"Extracted embedded image: {out_name}")
                    except Exception as e:
                        logger.warning(f"Failed to extract image {img_index} from page {page_index + 1}: {e}")
            
            doc.close()
            return extracted_images
        except Exception as e:
            logger.error(f"Error extracting embedded images from {pdf_path}: {e}")
            return []

# ==================== EXCEL WRITER ====================
class ExcelWriter:
    def __init__(self, excel_path):
        self.excel_path = excel_path
    
    def create_excel_with_data(self, data_list):
        df = pd.DataFrame(data_list)
        
        # Reorder columns
        columns = [
            "filename", "doc_num", "corner_of_section", "township", "range", "county",
            "image_present_bool", "image_present_flag",
            "township_options", "range_options", "county_options"
        ]
        df = df[columns]
        
        # Save to Excel
        df.to_excel(self.excel_path, index=False, engine="openpyxl")
        logger.info(f"Created Excel file with {len(df)} records: {self.excel_path}")
        return df
    
    def add_images_to_excel(self, screenshot_dir, embedded_img_dir):
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            # Find the filename column
            filename_col = 1
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header and "filename" in str(header).lower():
                    filename_col = col
                    break
            
            # Add headers for image columns
            screenshot_col = ws.max_column + 1
            embedded_img_col = ws.max_column + 2
            
            ws.cell(row=1, column=screenshot_col, value="PDF Screenshot")
            ws.cell(row=1, column=embedded_img_col, value="Embedded Image")
            
            # Set column widths
            ws.column_dimensions[get_column_letter(screenshot_col)].width = 40
            ws.column_dimensions[get_column_letter(embedded_img_col)].width = 40
            
            # Process each row
            for row in range(2, ws.max_row + 1):
                filename = ws.cell(row=row, column=filename_col).value
                if not filename:
                    continue
                
                filename_no_ext = os.path.splitext(str(filename))[0]
                
                # Add screenshot
                screenshot_added = self._add_image_to_cell(
                    ws, row, screenshot_col, screenshot_dir, filename_no_ext, "screenshot"
                )
                
                # Add embedded image
                embedded_added = self._add_image_to_cell(
                    ws, row, embedded_img_col, embedded_img_dir, filename_no_ext, "embedded"
                )
                
                # Adjust row height if any image was added
                if screenshot_added or embedded_added:
                    ws.row_dimensions[row].height = 150
            
            # Save the workbook
            output_path = self.excel_path.replace('.xlsx', '_with_images.xlsx')
            wb.save(output_path)
            logger.info(f"Saved Excel with images: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error adding images to Excel: {e}")
            return None
    
    def _add_image_to_cell(self, ws, row, col, img_dir, filename_no_ext, img_type):
        try:
            # Find matching image
            img_path = None
            for img_file in os.listdir(img_dir):
                img_name_no_ext = os.path.splitext(img_file)[0]
                
                if img_type == "screenshot":
                    # Match screenshot files
                    if (img_name_no_ext == filename_no_ext or 
                        img_name_no_ext.startswith(f"{filename_no_ext}_page")):
                        img_path = os.path.join(img_dir, img_file)
                        break
                else:
                    # Match embedded image files
                    if img_name_no_ext.startswith(f"{filename_no_ext}_embedded"):
                        img_path = os.path.join(img_dir, img_file)
                        break
            
            if img_path and os.path.exists(img_path):
                img = ExcelImage(img_path)
                img.width = 200
                img.height = 200
                cell = f"{get_column_letter(col)}{row}"
                ws.add_image(img, cell)
                logger.debug(f"Added {img_type} image for {filename_no_ext} in row {row}")
                return True
            else:
                logger.debug(f"No {img_type} image found for {filename_no_ext}")
                return False
                
        except Exception as e:
            logger.warning(f"Error adding {img_type} image for row {row}: {e}")
            return False

# ==================== MAIN PIPELINE ====================
class PDFToExcelPipeline:
    def __init__(self, pdf_dir, output_dir):
        self.pdf_dir = Path(pdf_dir)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Create subdirectories
        self.screenshot_dir = self.output_dir / "screenshots"
        self.embedded_img_dir = self.output_dir / "embedded_images"
        self.screenshot_dir.mkdir(parents=True, exist_ok=True)
        self.embedded_img_dir.mkdir(parents=True, exist_ok=True)
        
        # Initialize components
        self.form_extractor = PDFFormExtractor()
        self.screenshot_converter = PDFToImageConverter(self.screenshot_dir)
        self.embedded_extractor = EmbeddedImageExtractor(self.embedded_img_dir)
        
        # Excel output path
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.excel_path = self.output_dir / f"pdf_extraction_{timestamp}.xlsx"
        self.excel_writer = ExcelWriter(str(self.excel_path))
    
    def process_all_pdfs(self):
        # Find all PDFs
        pdf_files = list(self.pdf_dir.glob("*.pdf"))
        
        if not pdf_files:
            logger.error(f"No PDF files found in {self.pdf_dir}")
            return None
        
        logger.info(f"Found {len(pdf_files)} PDF files to process")
        
        all_results = []
        
        for pdf_path in pdf_files:
            logger.info(f"\n{'='*50}")
            logger.info(f"Processing: {pdf_path.name}")
            logger.info(f"{'='*50}")
            
            # Step 1: Extract form data
            form_data = self.form_extractor.process_pdf(str(pdf_path))
            
            if form_data:
                all_results.append(form_data)
                
                # Step 2: Convert PDF to screenshots
                screenshots = self.screenshot_converter.convert_pdf_to_images(str(pdf_path))
                logger.info(f"Created {len(screenshots)} screenshot(s)")
                
                # Step 3: Extract embedded images
                embedded_images = self.embedded_extractor.extract_embedded_images(str(pdf_path))
                logger.info(f"Extracted {len(embedded_images)} embedded image(s)")
        
        if not all_results:
            logger.error("No data extracted from PDFs")
            return None
        
        # Step 4: Create Excel with form data
        self.excel_writer.create_excel_with_data(all_results)
        
        # Step 5: Add images to Excel
        final_excel = self.excel_writer.add_images_to_excel(
            str(self.screenshot_dir),
            str(self.embedded_img_dir)
        )
        
        return final_excel

def main():
    parser = argparse.ArgumentParser(
        description="Complete pipeline to extract PDF data, images, and screenshots to Excel"
    )
    parser.add_argument(
        "--pdf-dir",
        default=Config.PDF_DIR,
        help="Directory containing PDF files"
    )
    parser.add_argument(
        "--output-dir",
        default=Config.OUTPUT_DIR,
        help="Output directory for results"
    )
    parser.add_argument(
        "--log-level",
        default="INFO",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        help="Logging level"
    )
    
    args = parser.parse_args()
    
    # Set logging level
    logger.setLevel(getattr(logging, args.log_level))
    
    # Create and run pipeline
    pipeline = PDFToExcelPipeline(args.pdf_dir, args.output_dir)
    
    logger.info("Starting PDF to Excel pipeline...")
    start_time = time.time()
    
    final_excel = pipeline.process_all_pdfs()
    
    elapsed_time = time.time() - start_time
    
    if final_excel:
        logger.info(f"\n{'='*60}")
        logger.info(f"Pipeline completed successfully in {elapsed_time:.2f} seconds")
        logger.info(f"Final Excel output: {final_excel}")
        logger.info(f"Screenshots saved in: {pipeline.screenshot_dir}")
        logger.info(f"Embedded images saved in: {pipeline.embedded_img_dir}")
        logger.info(f"{'='*60}")
    else:
        logger.error("Pipeline failed to complete")

if __name__ == "__main__":
    main()