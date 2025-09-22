import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# Specify the path to the existing Excel file and image folder
excel_file = r"C:\Users\sk23dg\Desktop\Labins_Workflow\PDF_Extractor\pdf_extractor_Draft\output3.0.xlsx"  # Replace with your Excel file path
image_folder = r"C:\Users\sk23dg\Desktop\Labins_Workflow\PDF_Extractor\test_pdfs\extracted_images"  # Replace with your image folder path
output_excel = "updated_excel_with_images.xlsx"

# Load the existing workbook and select the active sheet
wb = load_workbook(excel_file)
ws = wb.active

# Assume filenames are in column A (change to the correct column letter if different)
filename_column = 'A'
image_column = 'B'  # New column for images

# Set header for the new image column
ws[f'{image_column}1'] = "Image"

# Get list of image files (supporting common image formats)
image_extensions = ('.png', '.jpg', '.jpeg', '.bmp', '.gif')
image_files = {f: os.path.join(image_folder, f) for f in os.listdir(image_folder) if f.lower().endswith(image_extensions)}

# Set column width for images (adjust as needed)
ws.column_dimensions[image_column].width = 50

# Iterate through rows to match filenames and add images
row = 2  # Start from row 2 to skip headers
while ws[f'{filename_column}{row}'].value:
    filename = ws[f'{filename_column}{row}'].value
    if filename in image_files:
        # Load and insert image to the image column
        img_path = image_files[filename]
        img = Image(img_path)
        
        # Resize image to fit in cell (adjust dimensions as needed)
        img.width = 100  # Width in pixels
        img.height = 100  # Height in pixels
        
        # Add image to the cell in the image column
        cell = f'{image_column}{row}'
        ws.add_image(img, cell)
        
        # Adjust row height to accommodate image
        ws.row_dimensions[row].height = 80  # Height in points (adjust as needed)
    
    row += 1

# Save the updated workbook
wb.save(output_excel)
print(f"Updated Excel file '{output_excel}' created successfully.")