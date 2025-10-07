import openpyxl
from openpyxl.drawing.image import Image as XLImage
import pandas as pd
from PIL import Image
import io
import base64
import logging
from typing import List, Dict, Tuple, Optional
import os
import glob
from dataclasses import dataclass
import json
from concurrent.futures import ThreadPoolExecutor, as_completed
import hashlib
import time
from enum import Enum
from transformers import pipeline
import torch
import argparse
import shutil

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('content_moderation.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class ContentSeverity(Enum):
    SAFE = "safe"
    QUESTIONABLE = "questionable"
    EXPLICIT = "explicit"
    VIOLENT = "violent"
    DISTURBING = "disturbing"

@dataclass
class ModerationResult:
    is_explicit: bool
    confidence: float
    categories: List[str]
    severity: ContentSeverity
    details: Dict

class ImageExtractor:
    """Extract images from Excel cells"""
    
    @staticmethod
    def extract_images_from_xlsx(file_path: str) -> Dict[str, List[Tuple[int, int, bytes]]]:
        """
        Extract all images from Excel file
        Returns: Dict mapping sheet names to list of (row, col, image_bytes)
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True, keep_vba=True)
            images_data = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_images = []
                
                # Handle embedded images
                if hasattr(sheet, '_images') and sheet._images:
                    for image in sheet._images:
                        # Get image position
                        row = image.anchor._from.row + 1  # 1-based
                        col = image.anchor._from.col + 1  # 1-based
                        
                        # Extract image data as bytes (fix for buffer API error)
                        try:
                            # Reset the stream position if needed
                            if hasattr(image.ref, 'seek'):
                                image.ref.seek(0)
                            img_bytes = image._data()  # This returns raw bytes
                        except Exception as e:
                            logger.warning(f"Failed to extract embedded image data: {e}")
                            continue
                        
                        if img_bytes:
                            sheet_images.append((row, col, img_bytes))
                        else:
                            logger.warning(f"No data extracted for embedded image at row {row}, col {col}")
                
                # Handle images in cells (base64 encoded or binary)
                for row_obj in sheet.iter_rows():
                    for cell in row_obj:
                        if cell.value and isinstance(cell.value, bytes):
                            sheet_images.append((cell.row, cell.column, cell.value))
                        elif cell.value and isinstance(cell.value, str):
                            # Check if it's base64 encoded image
                            if cell.value.startswith('data:image'):
                                try:
                                    # Extract base64 data
                                    base64_data = cell.value.split(',')[1]
                                    img_bytes = base64.b64decode(base64_data)
                                    sheet_images.append((cell.row, cell.column, img_bytes))
                                except Exception as e:
                                    logger.warning(f"Failed to decode base64 image at {cell.coordinate}: {e}")
                
                if sheet_images:
                    images_data[sheet_name] = sheet_images
            
            return images_data
            
        except Exception as e:
            logger.error(f"Error extracting images from {file_path}: {e}")
            raise

class ContentModerationAPI:
    """Local content moderation using pre-trained ML model (no API keys or external calls required)"""
    
    def __init__(self):
        self.rate_limit_delay = 0.1  # Minimal delay for local processing
        self.device = "cuda" if torch.cuda.is_available() else "cpu"
        logger.info(f"Using device: {self.device}")
        
        # Load the NSFW detection pipeline (downloads model on first run, then runs entirely locally)
        self.detector = pipeline(
            "image-classification", 
            model="Falconsai/nsfw_image_detection",
            device=self.device if self.device == "cuda" else -1  # -1 for CPU
        )
        logger.info("NSFW detection model loaded successfully")
    
    def analyze_image(self, image_bytes: bytes) -> ModerationResult:
        """
        Analyze image for explicit content using local model
        """
        try:
            return self._local_nsfw_detection(image_bytes)
                
        except Exception as e:
            logger.error(f"Local model analysis failed: {e}")
            # Return safe by default on failure
            return ModerationResult(
                is_explicit=False,
                confidence=0.0,
                categories=[],
                severity=ContentSeverity.SAFE,
                details={"error": str(e)}
            )
    
    def _local_nsfw_detection(self, image_bytes: bytes) -> ModerationResult:
        """Local NSFW detection using pre-trained model"""
        # Convert bytes to PIL Image
        img = Image.open(io.BytesIO(image_bytes))
        if img.mode != 'RGB':
            img = img.convert('RGB')  # Ensure RGB for model compatibility
        
        time.sleep(self.rate_limit_delay)  # Minimal delay
        
        # Run detection (entirely local)
        prediction = self.detector(img)
        
        # Parse prediction (model outputs 'nsfw' or 'normal')
        top_prediction = prediction[0]
        top_label = top_prediction['label'].lower()
        confidence = top_prediction['score']
        
        is_explicit = top_label == 'nsfw' and confidence > 0.5  # Adjustable threshold
        categories = ['adult'] if top_label == 'nsfw' else []
        
        severity = ContentSeverity.SAFE
        if is_explicit:
            severity = ContentSeverity.EXPLICIT
        elif top_label == 'nsfw' and confidence > 0.3:  # Lower threshold for questionable
            severity = ContentSeverity.QUESTIONABLE
        
        return ModerationResult(
            is_explicit=is_explicit,
            confidence=confidence,
            categories=categories,
            severity=severity,
            details={
                "top_label": top_label,
                "full_prediction": prediction
            }
        )

class ExcelContentModerator:
    """Main class for moderating Excel content"""
    
    def __init__(self, 
                 sensitivity_threshold: float = 0.5,
                 backup_original: bool = True):
        self.extractor = ImageExtractor()
        self.moderator = ContentModerationAPI()  # No API key needed
        self.sensitivity_threshold = sensitivity_threshold
        self.backup_original = backup_original
        self.processed_images_cache = {}  # Cache to avoid reprocessing same images
        
    def process_excel_file(self, input_path: str, output_path: str = None,
                          target_columns: List[str] = None) -> Dict:
        """
        Process a single Excel file and remove rows with explicit content
        """
        if not output_path:
            base, ext = os.path.splitext(input_path)
            output_path = f"{base}_cleaned{ext}"
            
        # Create backup if requested
        if self.backup_original:
            backup_path = f"{input_path}.backup"
            shutil.copy2(input_path, backup_path)
            logger.info(f"Created backup at {backup_path}")
            
        results = {
            'file': input_path,
            'total_rows': 0,
            'removed_rows': 0,
            'processed_images': 0,
            'explicit_images': 0,
            'errors': [],
            'removed_row_details': []
        }
        
        try:
            # Extract images
            logger.info(f"Extracting images from {input_path}")
            images_data = self.extractor.extract_images_from_xlsx(input_path)
            
            # Load Excel with pandas for easier row manipulation
            excel_file = pd.ExcelFile(input_path)
            output_sheets = {}
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                results['total_rows'] += len(df)
                
                if sheet_name in images_data:
                    rows_to_remove = set()
                    
                    # Process images with threading for performance
                    with ThreadPoolExecutor(max_workers=5) as executor:
                        future_to_image = {}
                        
                        for row_idx, col_idx, img_bytes in images_data[sheet_name]:
                            # Validate img_bytes is bytes
                            if not isinstance(img_bytes, bytes):
                                logger.error(f"Invalid image data type at row {row_idx}, col {col_idx}: {type(img_bytes)}")
                                continue
                            
                            # Create hash for caching
                            img_hash = hashlib.md5(img_bytes).hexdigest()
                            
                            if img_hash in self.processed_images_cache:
                                # Use cached result
                                moderation_result = self.processed_images_cache[img_hash]
                            else:
                                # Submit for processing
                                future = executor.submit(self.moderator.analyze_image, img_bytes)
                                future_to_image[future] = (row_idx, col_idx, img_hash)
                        
                        # Collect results
                        for future in as_completed(future_to_image):
                            row_idx, col_idx, img_hash = future_to_image[future]
                            
                            try:
                                moderation_result = future.result(timeout=30)
                                self.processed_images_cache[img_hash] = moderation_result
                                
                                results['processed_images'] += 1
                                
                                if moderation_result.is_explicit or \
                                   (moderation_result.confidence > self.sensitivity_threshold and 
                                    moderation_result.severity != ContentSeverity.SAFE):
                                    
                                    rows_to_remove.add(row_idx - 1)  # Adjust for 0-based pandas index
                                    results['explicit_images'] += 1
                                    results['removed_row_details'].append({
                                        'sheet': sheet_name,
                                        'row': row_idx,
                                        'column': col_idx,
                                        'severity': moderation_result.severity.value,
                                        'confidence': moderation_result.confidence,
                                        'categories': moderation_result.categories
                                    })
                                    
                                    logger.warning(f"Explicit content found in {sheet_name} at row {row_idx}, col {col_idx}")
                                    
                            except Exception as e:
                                logger.error(f"Error processing image at {sheet_name}[{row_idx},{col_idx}]: {e}")
                                results['errors'].append(str(e))
                    
                    # Remove flagged rows
                    if rows_to_remove:
                        df_cleaned = df.drop(index=list(rows_to_remove))
                        results['removed_rows'] += len(rows_to_remove)
                        logger.info(f"Removed {len(rows_to_remove)} rows from {sheet_name}")
                    else:
                        df_cleaned = df
                else:
                    df_cleaned = df
                    
                output_sheets[sheet_name] = df_cleaned
            
            # Save cleaned Excel file
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, df in output_sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
            logger.info(f"Cleaned file saved to {output_path}")
            
            # Save detailed report for this file
            report_path = f"{output_path}_report.json"
            with open(report_path, 'w') as f:
                json.dump(results, f, indent=2)
            logger.info(f"Report saved to {report_path}")
            
            return results
            
        except Exception as e:
            logger.error(f"Critical error processing file: {e}")
            results['errors'].append(f"Critical error: {e}")
            raise

    def process_multiple_files(self, input_paths: List[str], output_dir: str = None) -> Dict:
        """
        Process multiple Excel files and aggregate results.
        :param input_paths: List of input Excel file paths
        :param output_dir: Optional directory for outputs (defaults to input dir)
        :return: Aggregated results across all files
        """
        if not output_dir:
            output_dir = os.path.dirname(input_paths[0]) if input_paths else '.'

        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)

        aggregated_results = {
            'files_processed': 0,
            'total_rows': 0,
            'total_removed_rows': 0,
            'total_processed_images': 0,
            'total_explicit_images': 0,
            'errors': [],
            'file_results': {}
        }

        for input_path in input_paths:
            if not os.path.exists(input_path):
                logger.warning(f"File not found: {input_path}")
                aggregated_results['errors'].append(f"File not found: {input_path}")
                continue

            try:
                # Determine output path for this file
                base_name = os.path.basename(input_path)
                base, ext = os.path.splitext(base_name)
                output_path = os.path.join(output_dir, f"{base}_cleaned{ext}")

                # Process single file
                file_results = self.process_excel_file(input_path, output_path)
                aggregated_results['file_results'][input_path] = file_results

                # Aggregate
                aggregated_results['files_processed'] += 1
                aggregated_results['total_rows'] += file_results['total_rows']
                aggregated_results['total_removed_rows'] += file_results['removed_rows']
                aggregated_results['total_processed_images'] += file_results['processed_images']
                aggregated_results['total_explicit_images'] += file_results['explicit_images']
                aggregated_results['errors'].extend(file_results['errors'])

                logger.info(f"Completed processing: {input_path}")

            except Exception as e:
                logger.error(f"Failed to process {input_path}: {e}")
                aggregated_results['errors'].append(f"Failed to process {input_path}: {e}")

        # Save aggregated report
        aggregated_report_path = os.path.join(output_dir, 'aggregated_moderation_report.json')
        with open(aggregated_report_path, 'w') as f:
            json.dump(aggregated_results, f, indent=2)
        logger.info(f"Aggregated report saved to {aggregated_report_path}")

        return aggregated_results

# Main execution with CLI arguments
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Excel Content Moderator: Detect and remove explicit images from Excel files using local ML model.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Process all .xlsx files in a directory
  python excel_moderator.py --input_dir /path/to/input/dir --output_dir /path/to/output/dir

  # Process specific files
  python excel_moderator.py /path/to/file1.xlsx /path/to/file2.xlsx --output_dir /path/to/output/dir

  # Single file with custom sensitivity
  python excel_moderator.py /path/to/single_file.xlsx --output_dir /path/to/output --sensitivity 0.7 --no-backup
        """
    )
    
    parser.add_argument(
        '--input_dir',
        type=str,
        help='Directory containing Excel files to process (uses *.xlsx pattern)'
    )
    
    parser.add_argument(
        'input_files',
        nargs='*',
        type=str,
        help='Specific Excel file paths to process (positional arguments)'
    )
    
    parser.add_argument(
        '--output_dir',
        type=str,
        default='.',
        help='Output directory for cleaned files and reports (default: current directory)'
    )
    
    parser.add_argument(
        '--sensitivity',
        type=float,
        default=0.5,
        help='Sensitivity threshold for flagging content (0.0 to 1.0, default: 0.5)'
    )
    
    parser.add_argument(
        '--no-backup',
        action='store_true',
        help='Disable creating backups of original files'
    )
    
    args = parser.parse_args()
    
    # Collect input paths
    input_paths = []
    
    if args.input_dir:
        if input_paths:  # If both dir and files provided, warn
            logger.warning("Both --input_dir and input_files provided. Using input_files only.")
        else:
            input_paths = glob.glob(os.path.join(args.input_dir, "*.xlsx"))
            if not input_paths:
                logger.error(f"No .xlsx files found in {args.input_dir}")
                exit(1)
    
    if args.input_files:
        input_paths.extend(args.input_files)
    
    if not input_paths:
        parser.error("Must provide either --input_dir or at least one input file path.")
    
    # Initialize moderator
    moderator = ExcelContentModerator(
        sensitivity_threshold=args.sensitivity,
        backup_original=not args.no_backup
    )
    
    try:
        results = moderator.process_multiple_files(
            input_paths=input_paths,
            output_dir=args.output_dir
        )
        
        print("\n=== Multi-File Processing Complete ===")
        print(f"Files processed: {results['files_processed']}")
        print(f"Total rows processed: {results['total_rows']}")
        print(f"Total rows removed: {results['total_removed_rows']}")
        print(f"Total images analyzed: {results['total_processed_images']}")
        print(f"Total explicit images found: {results['total_explicit_images']}")
        
        if results['errors']:
            print(f"\nErrors encountered: {len(results['errors'])}")
            for error in results['errors'][:5]:  # Show first 5 errors
                print(f"  - {error}")
                
    except Exception as e:
        logger.error(f"Failed to process files: {e}")
        raise