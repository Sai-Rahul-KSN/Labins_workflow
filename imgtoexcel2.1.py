import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# Specify the paths
excel_file = r"C:\Users\sk23dg\Desktop\Labins_Workflow\PDF_Extractor\pdf_extractor_Draft\output3.0.xlsx"  # Replace with your Excel file path
image_folder = r"C:\Users\sk23dg\Desktop\Labins_Workflow\PDF_Extractor\test_pdfs\extracted_images"  # Replace with your image folder path
output_excel = "updated_excel_with_images.xlsx"

# Validate paths
if not os.path.exists(excel_file):
    print(f"Error: Excel file '{excel_file}' does not exist.")
    exit()
if not os.path.exists(image_folder):
    print(f"Error: Image folder '{image_folder}' does not exist.")
    exit()

# Load the existing workbook
try:
    wb = load_workbook(excel_file)
except Exception as e:
    print(f"Error loading Excel file: {e}")
    exit()

ws = wb.active
print(f"Active sheet: {ws.title}")

# Try to detect which column has the filenames by header name (e.g. "File Name").
# Fallback to column A if no matching header is found.
filename_column = 'A'
filename_col_index = 1
image_column = 'B'  # will be updated below

# Look for common header variants in the first row.
possible_header_found = False
possible_headers = ["file name", "filename", "file_name", "file"]
for col_idx in range(1, ws.max_column + 1):
    header_cell = ws.cell(row=1, column=col_idx).value
    if header_cell:
        header_text = str(header_cell).strip().lower()
        # Match exact common variants or headers that contain both words
        if header_text in possible_headers or ("file" in header_text and "name" in header_text) or "filename" in header_text:
            from openpyxl.utils import get_column_letter
            filename_col_index = col_idx
            filename_column = get_column_letter(col_idx)
            image_column = get_column_letter(col_idx + 1)
            possible_header_found = True
            print(f"Detected filename header '{header_cell}' in column {filename_column}")
            break

if not possible_header_found:
    # default behavior
    filename_column = 'A'
    filename_col_index = 1
    image_column = 'B'
    print("No filename header detected; defaulting to column A for filenames and B for images.")

# Set header for the image column
ws[f'{image_column}1'] = "Image"

# Get list of image files
image_extensions = ('.png', '.jpg', '.jpeg', '.bmp', '.gif')
image_files = {f: os.path.join(image_folder, f) for f in os.listdir(image_folder) if f.lower().endswith(image_extensions)}
print(f"Found {len(image_files)} image files: {list(image_files.keys())}")

# Set column width for images
ws.column_dimensions[image_column].width = 50

# Iterate through rows to match filenames and add images
row = 2
images_added = 0
while ws[f'{filename_column}{row}'].value:
    filename = ws[f'{filename_column}{row}'].value
    # Clean filename to handle extensions or spaces
    filename = str(filename).strip()
    # Try matching with and without extension
    filename_no_ext = os.path.splitext(filename)[0]
    matched_file = None
    for img_file in image_files:
        if img_file == filename or os.path.splitext(img_file)[0] == filename_no_ext:
            matched_file = img_file
            break
    
    if matched_file:
        try:
            img_path = image_files[matched_file]
            img = Image(img_path)
            # Resize image
            img.width = 100
            img.height = 100
            # Add image to cell
            cell = f'{image_column}{row}'
            ws.add_image(img, cell)
            # Adjust row height
            ws.row_dimensions[row].height = 80
            images_added += 1
            print(f"Added image '{matched_file}' to row {row}")
        except Exception as e:
            print(f"Error adding image '{matched_file}' to row {row}: {e}")
    else:
        print(f"No matching image found for filename '{filename}' in row {row}")
    
    row += 1

# Save the workbook
try:
    wb.save(output_excel)
    print(f"Excel file '{output_excel}' saved successfully. {images_added} images added.")
except Exception as e:
    print(f"Error saving Excel file: {e}")