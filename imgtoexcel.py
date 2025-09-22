import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# Specify the folder containing images
image_folder = r"C:\Users\sk23dg\Desktop\Labins_Workflow\PDF_Extractor\test_pdfs\extracted_images"  # Replace with your folder path
output_excel = r"output_images.xlsx"

# Create a new workbook and select the active sheet
wb = Workbook()
ws = wb.active
ws.title = "Images"

# Set column headers
ws['A1'] = "File Name"
ws['B1'] = "Image"

# Get list of image files (supporting common image formats)
image_extensions = ('.png', '.jpg', '.jpeg', '.bmp', '.gif')
image_files = [f for f in os.listdir(image_folder) if f.lower().endswith(image_extensions)]

# Set column widths (adjust as needed)
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 50

# Iterate through images and add to Excel
row = 2  # Start from row 2 to skip headers
for image_file in image_files:
    # Write file name to column A
    ws[f'A{row}'] = image_file
    
    # Load and insert image to column B
    img_path = os.path.join(image_folder, image_file)
    img = Image(img_path)
    
    # Resize image to fit in cell (adjust dimensions as needed)
    img.width = 100  # Width in pixels
    img.height = 100  # Height in pixels
    
    # Add image to the cell in column B
    cell = f'B{row}'
    ws.add_image(img, cell)
    
    # Adjust row height to accommodate image
    ws.row_dimensions[row].height = 80  # Height in points (adjust as needed)
    
    row += 1

# Save the workbook
wb.save(output_excel)
print(f"Excel file '{output_excel}' created successfully.")